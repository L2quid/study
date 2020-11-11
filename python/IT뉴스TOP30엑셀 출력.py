from openpyxl import Workbook
from bs4 import BeautifulSoup
import urllib.request
import datetime
import os
date= datetime.datetime.today().strftime('%Y%m%d')
url = 'https://news.naver.com/main/ranking/popularDay.nhn?rankingType=popular_day&sectionId=105&date='+date
with urllib.request.urlopen(url) as fs :
    soup = BeautifulSoup(fs.read().decode(fs.headers.get_content_charset()), 'html.parser')
    items = soup.find_all('div', {'class' : 'ranking_headline'})
    links=soup.find_all('a',{'class':'nclicks(rnk.sci)'})
desktopPath = os.path.expanduser('~')
filePath = desktopPath + '\뉴스'
wb = Workbook()
ws = wb.active
ws.column_dimensions['B'].width = 60
ws.column_dimensions['C'].width = 12
for i in range (30):
    ws.cell(row=i+1,column=1).value=str(i+1)+'위'
    ws.cell(row=i+1,column=2).value=items[i].get_text(strip = True)
    ws.cell(row=i+1,column=3).value=str('링크로 이동')
    ws.cell(row=i+1,column=3).hyperlink=str('https://news.naver.com')+links[2*i].attrs['href']
wb.save(filePath+date+".xlsx")
